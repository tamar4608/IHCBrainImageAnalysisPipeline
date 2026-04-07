# nb05_coloc_report.py
# ============================================================
# Notebook 5 — Co-localization & Report
# Match Orexin+ cells with cFOS+ cells, export Excel + summary.
# ============================================================
"""
Workflow:
  1. Load GFP (Orexin) and Cy5 (cFOS) centroids per slide.
  2. Adjust distance threshold (µm) for co-localization.
  3. Preview overlay: GFP=green, cFOS=red, double-positive=yellow.
  4. Pull region-level counts from atlas annotations.
  5. Export Excel workbook (multi-sheet) + print summary.

Output: data/results/coloc_results.xlsx
"""

import sys
import csv
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
from utils import (
    load_config, SessionStore,
    centroids_from_masks, match_cells,
    area_px_to_mm2,
)

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider

# ── openpyxl (Excel export) ───────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    print("⚠️  openpyxl not installed — Excel export disabled.")
    EXCEL_OK = False

# ── Config & session ──────────────────────────────────────────────────────────
cfg      = load_config()
session  = SessionStore(cfg["paths"]["session"])
PROC_DIR = Path(cfg["paths"]["processed"])
RES_DIR  = Path(cfg["paths"]["results"])
RES_DIR.mkdir(parents=True, exist_ok=True)

RESOLUTION_UM = 25.0                               # µm per pixel (10x scan)
COLOC_UM_DEFAULT = cfg["coloc"]["distance_threshold_um"]
COLOC_PX_DEFAULT = COLOC_UM_DEFAULT / RESOLUTION_UM

# ── Approved slides ───────────────────────────────────────────────────────────
selections = session.load("slice_selections", default={})
det_results = session.load("detection_results", default={})

ready_slides = [
    Path(v["slide"])
    for k, v in selections.items()
    if v["decision"] == "include"
    and det_results.get(k, {}).get("flag") == "ok"
]

if not ready_slides:
    # Fallback: all detected slides
    ready_slides = [
        Path(v["slide"])
        for k, v in selections.items()
        if v["decision"] == "include"
        and (PROC_DIR / Path(v["slide"]).stem / "GFP_masks.npy").exists()
    ]

print(f"Slides for co-localization: {len(ready_slides)}")
if not ready_slides:
    print("No slides ready. Complete nb04 first.")
    sys.exit(0)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_centroids(slide_path: Path, ch: str) -> np.ndarray:
    """Load centroid CSV → (N, 2) float array. Returns empty array if missing."""
    p = PROC_DIR / slide_path.stem / f"{ch}_centroids.csv"
    if not p.exists():
        return np.empty((0, 2), dtype=float)
    with open(p) as f:
        reader = csv.DictReader(f)
        rows   = [(float(r["row_px"]), float(r["col_px"])) for r in reader]
    return np.array(rows) if rows else np.empty((0, 2), dtype=float)


def _load_masks(slide_path: Path, ch: str) -> np.ndarray | None:
    p = PROC_DIR / slide_path.stem / f"{ch}_masks.npy"
    return np.load(p) if p.exists() else None


def compute_coloc(slide_path: Path, dist_px: float) -> dict:
    """
    Returns:
      orexin_n    — Orexin+ count
      cfos_n      — cFOS+ count
      dapi_n      — DAPI count
      coloc_n     — double-positive (Orexin+ & cFOS+) count
      coloc_pct   — % of Orexin+ that are also cFOS+
      orexin_xy   — (N,2) centroids
      cfos_xy     — (M,2) centroids
      coloc_pairs — (K,2) index pairs
    """
    orexin_xy = _load_centroids(slide_path, "GFP")
    cfos_xy   = _load_centroids(slide_path, "Cy5")
    dapi_xy   = _load_centroids(slide_path, "DAPI")

    pairs   = match_cells(orexin_xy, cfos_xy, max_dist_px=dist_px)
    coloc_n = len(pairs)
    pct     = (coloc_n / len(orexin_xy) * 100) if len(orexin_xy) > 0 else 0.0

    return {
        "orexin_n":    len(orexin_xy),
        "cfos_n":      len(cfos_xy),
        "dapi_n":      len(dapi_xy),
        "coloc_n":     coloc_n,
        "coloc_pct":   round(pct, 2),
        "orexin_xy":   orexin_xy,
        "cfos_xy":     cfos_xy,
        "coloc_pairs": pairs,
    }


def make_coloc_overlay(slide_path: Path, result: dict) -> np.ndarray | None:
    """Green=Orexin, Red=cFOS, Yellow=double-positive overlay on DAPI."""
    dapi_p = PROC_DIR / slide_path.stem / "DAPI.npy"
    if not dapi_p.exists():
        return None

    dapi = np.load(dapi_p)
    rgb  = np.stack([dapi * 0.3, dapi * 0.3, dapi * 0.5], axis=-1)  # dim blue-grey BG
    H, W = dapi.shape

    def _dot(arr, color, r=4):
        for row, col in arr:
            rr, cc = int(row), int(col)
            for dr in range(-r, r + 1):
                for dc in range(-r, r + 1):
                    if dr*dr + dc*dc <= r*r:
                        nr, nc = rr + dr, cc + dc
                        if 0 <= nr < H and 0 <= nc < W:
                            rgb[nr, nc] = color

    _dot(result["orexin_xy"], [0, 1, 0])   # green
    _dot(result["cfos_xy"],   [1, 0, 0])   # red

    # Highlight double-positives in yellow
    coloc_orexin = result["orexin_xy"][result["coloc_pairs"][:, 0]] if len(result["coloc_pairs"]) else np.empty((0, 2))
    _dot(coloc_orexin, [1, 1, 0])          # yellow

    return np.clip(rgb, 0, 1)


# ── Interactive distance threshold ────────────────────────────────────────────
print("\nAdjust co-localization distance threshold.")
dist_state = [COLOC_PX_DEFAULT]

if ready_slides:
    rep = ready_slides[0]
    test_result = compute_coloc(rep, dist_state[0])

    fig, ax = plt.subplots(figsize=(7, 7))
    plt.subplots_adjust(bottom=0.2)

    ov = make_coloc_overlay(rep, test_result)
    if ov is not None:
        im = ax.imshow(ov)
    ax.set_title(f"{rep.name}\nOrexin={test_result['orexin_n']}  "
                 f"cFOS={test_result['cfos_n']}  "
                 f"Coloc={test_result['coloc_n']} ({test_result['coloc_pct']}%)")
    ax.axis("off")

    ax_dist = plt.axes([0.15, 0.08, 0.65, 0.04])
    s_dist  = Slider(ax_dist, "Distance (px)", 1, 50,
                     valinit=COLOC_PX_DEFAULT, valstep=0.5)

    def _update_dist(val):
        dist_state[0] = float(s_dist.val)
        r = compute_coloc(rep, dist_state[0])
        ax.set_title(f"{rep.name}\nOrexin={r['orexin_n']}  "
                     f"cFOS={r['cfos_n']}  "
                     f"Coloc={r['coloc_n']} ({r['coloc_pct']}%)")
        ov2 = make_coloc_overlay(rep, r)
        if ov2 is not None and ov is not None:
            im.set_data(ov2)
        fig.canvas.draw_idle()

    s_dist.on_changed(_update_dist)
    fig.suptitle("Green=Orexin  Red=cFOS  Yellow=double-positive\nClose window when satisfied")
    plt.show()
    plt.close("all")

final_dist_px = dist_state[0]
final_dist_um = final_dist_px * RESOLUTION_UM
print(f"Final distance threshold: {final_dist_px:.1f} px  ({final_dist_um:.1f} µm)")


# ── Compute all results ───────────────────────────────────────────────────────
all_rows = []

for slide in ready_slides:
    result = compute_coloc(slide, final_dist_px)
    row = {
        "slide":        slide.name,
        "hemisphere":   "A",          # both hemispheres are in one TIF; future: split
        "DAPI_count":   result["dapi_n"],
        "Orexin_count": result["orexin_n"],
        "cFOS_count":   result["cfos_n"],
        "Coloc_count":  result["coloc_n"],
        "Coloc_pct":    result["coloc_pct"],
    }

    # Region area from annotation mask (if available)
    ann_p = PROC_DIR / slide.stem / "annotation.npy"
    if ann_p.exists():
        ann = np.load(ann_p)
        region_px = int(np.sum(ann > 0))
        row["region_area_mm2"] = round(area_px_to_mm2(region_px, RESOLUTION_UM), 4)
        row["DAPI_density"]    = round(result["dapi_n"] / (row["region_area_mm2"] + 1e-9), 2)
    else:
        row["region_area_mm2"] = None
        row["DAPI_density"]    = None

    all_rows.append(row)
    print(f"  {slide.name}:  Orexin={result['orexin_n']}  "
          f"cFOS={result['cfos_n']}  Coloc={result['coloc_n']} ({result['coloc_pct']}%)")


# ── Excel export ──────────────────────────────────────────────────────────────

def _hdr_style(cell, wb):
    cell.font      = Font(bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center")


def export_excel(rows: list[dict], out_path: Path) -> None:
    wb = openpyxl.Workbook()

    def _add_sheet(title: str, headers: list[str], data: list[list]) -> None:
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for cell in ws[1]:
            _hdr_style(cell, wb)
        for row in data:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                12, max(len(str(c.value or "")) for c in col) + 2)

    # Sheet 1 — Summary
    _add_sheet(
        "Summary",
        ["Slide", "Hemisphere", "DAPI", "Orexin+", "cFOS+",
         "Orexin+cFOS+", "% Orexin+ active", "Region area (mm²)", "DAPI density (cells/mm²)"],
        [[r["slide"], r["hemisphere"], r["DAPI_count"], r["Orexin_count"],
          r["cFOS_count"], r["Coloc_count"], r["Coloc_pct"],
          r["region_area_mm2"], r["DAPI_density"]]
         for r in rows],
    )

    # Sheet 2 — DAPI counts
    _add_sheet("DAPI Counts",
               ["Slide", "Hemisphere", "DAPI Count"],
               [[r["slide"], r["hemisphere"], r["DAPI_count"]] for r in rows])

    # Sheet 3 — Orexin counts
    _add_sheet("Orexin Counts",
               ["Slide", "Hemisphere", "Orexin+ Count"],
               [[r["slide"], r["hemisphere"], r["Orexin_count"]] for r in rows])

    # Sheet 4 — cFOS counts
    _add_sheet("cFOS Counts",
               ["Slide", "Hemisphere", "cFOS+ Count"],
               [[r["slide"], r["hemisphere"], r["cFOS_count"]] for r in rows])

    # Sheet 5 — Co-localization
    _add_sheet("Co-localization",
               ["Slide", "Hemisphere", "Orexin+", "Coloc (Orexin+cFOS+)", "% Orexin+ active"],
               [[r["slide"], r["hemisphere"], r["Orexin_count"],
                 r["Coloc_count"], r["Coloc_pct"]] for r in rows])

    # Sheet 6 — Hemisphere breakdown (placeholder — A and B split TBD)
    _add_sheet("Hemisphere Breakdown",
               ["Slide", "Hemisphere A note"],
               [[r["slide"], "Both hemispheres in single TIF — PI to decide L/R assignment"]
                for r in rows])

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Metadata
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.append(["Generated",       datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_meta.append(["Distance thr (µm)", round(final_dist_um, 2)])
    ws_meta.append(["Resolution (µm/px)", RESOLUTION_UM])
    ws_meta.append(["Total slides",    len(rows)])

    wb.save(out_path)
    print(f"\n  ✅  Excel saved: {out_path}")


if EXCEL_OK:
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = RES_DIR / f"coloc_results_{ts}.xlsx"
    export_excel(all_rows, out_xlsx)
else:
    # Fallback: CSV
    out_csv = RES_DIR / "coloc_results.csv"
    import csv as _csv
    with open(out_csv, "w", newline="") as f:
        writer = _csv.DictWriter(f, fieldnames=list(all_rows[0].keys()) if all_rows else [])
        writer.writeheader()
        for r in all_rows:
            # strip non-serialisable arrays
            writer.writerow({k: v for k, v in r.items() if not isinstance(v, np.ndarray)})
    print(f"\n  ✅  CSV saved: {out_csv}")

# ── Print summary ─────────────────────────────────────────────────────────────
if all_rows:
    total_orexin = sum(r["Orexin_count"] for r in all_rows)
    total_coloc  = sum(r["Coloc_count"]  for r in all_rows)
    overall_pct  = round(total_coloc / total_orexin * 100, 2) if total_orexin else 0.0

    print("\n" + "=" * 55)
    print("  FINAL SUMMARY")
    print("=" * 55)
    print(f"  Slides analysed:          {len(all_rows)}")
    print(f"  Total Orexin+ cells:      {total_orexin}")
    print(f"  Total cFOS+ cells:        {sum(r['cFOS_count']  for r in all_rows)}")
    print(f"  Orexin+/cFOS+ (coloc):   {total_coloc}")
    print(f"  Overall % Orexin active:  {overall_pct}%")
    print("=" * 55)

print("\n🧠  Pipeline complete — please send results to the PI for review.")
